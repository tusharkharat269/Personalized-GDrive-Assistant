import io
import json
from typing import Callable

from bs4 import BeautifulSoup
from docx import Document as DocxDocument
from openpyxl import load_workbook
from pypdf import PdfReader

from app.core.exceptions import AppException
from app.core.logging import logger

# Google Workspace file export mappings
GOOGLE_DOC_MIME = "application/vnd.google-apps.document"
GOOGLE_SHEET_MIME = "application/vnd.google-apps.spreadsheet"
GOOGLE_SLIDE_MIME = "application/vnd.google-apps.presentation"

GOOGLE_EXPORT_MAP: dict[str, str] = {
    GOOGLE_DOC_MIME: "text/plain",
    GOOGLE_SHEET_MIME: "text/csv",
    GOOGLE_SLIDE_MIME: "text/plain",
}


def isGoogleWorkspaceMime(mimeType: str | None) -> bool:
    return bool(mimeType and mimeType.startswith("application/vnd.google-apps."))


def exportMimeFor(mimeType: str) -> str | None:
    return GOOGLE_EXPORT_MAP.get(mimeType)


class DocumentLoader:
    """Parses bytes of various document types into plain text."""

    def parse(self, content: bytes, mimeType: str, fileName: str = "") -> str:
        handler = self._pickHandler(mimeType, fileName)
        try:
            text = handler(content)
        except Exception as e:
            logger.error("document_parse_failed", mimeType=mimeType, fileName=fileName, error=str(e))
            raise AppException(400, "DOC_PARSE_FAILED", f"Failed to parse '{fileName}': {e}")
        return self._normalize(text)

    def _pickHandler(self, mimeType: str, fileName: str) -> Callable[[bytes], str]:
        mt = (mimeType or "").lower()
        lower = (fileName or "").lower()

        if mt == "application/pdf" or lower.endswith(".pdf"):
            return self._parsePdf
        if mt == "application/vnd.openxmlformats-officedocument.wordprocessingml.document" or lower.endswith(".docx"):
            return self._parseDocx
        if mt == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" or lower.endswith(".xlsx"):
            return self._parseXlsx
        if mt in ("text/html", "application/xhtml+xml") or lower.endswith((".html", ".htm")):
            return self._parseHtml
        if mt == "application/json" or lower.endswith(".json"):
            return self._parseJson
        # csv / md / plain / fallback
        return self._parseText

    # ---------- handlers ----------

    def _parsePdf(self, content: bytes) -> str:
        reader = PdfReader(io.BytesIO(content))
        parts: list[str] = []
        for i, page in enumerate(reader.pages):
            try:
                parts.append(page.extract_text() or "")
            except Exception as e:
                logger.warning("pdf_page_parse_failed", pageIndex=i, error=str(e))
        return "\n\n".join(parts)

    def _parseDocx(self, content: bytes) -> str:
        doc = DocxDocument(io.BytesIO(content))
        return "\n".join(p.text for p in doc.paragraphs if p.text)

    def _parseXlsx(self, content: bytes) -> str:
        wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
        lines: list[str] = []
        for ws in wb.worksheets:
            lines.append(f"# Sheet: {ws.title}")
            for row in ws.iter_rows(values_only=True):
                cells = ["" if v is None else str(v) for v in row]
                if any(cells):
                    lines.append(", ".join(cells))
        return "\n".join(lines)

    def _parseHtml(self, content: bytes) -> str:
        soup = BeautifulSoup(content, "html.parser")
        for tag in soup(["script", "style", "noscript"]):
            tag.decompose()
        return soup.get_text(separator="\n", strip=True)

    def _parseJson(self, content: bytes) -> str:
        try:
            data = json.loads(content.decode("utf-8", errors="ignore"))
            return json.dumps(data, indent=2, ensure_ascii=False)
        except Exception:
            return content.decode("utf-8", errors="ignore")

    def _parseText(self, content: bytes) -> str:
        return content.decode("utf-8", errors="ignore")

    def _normalize(self, text: str) -> str:
        return "\n".join(line.rstrip() for line in (text or "").splitlines() if line.strip())
